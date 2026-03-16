# Flask Web Application for Placement Cell Online Test Platform
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_from_directory
from functools import wraps
import os
from datetime import timedelta
import requests
import hashlib
import secrets
import smtplib
import time
from email.message import EmailMessage

# Supabase client wrapper: try the official Python client first; if it fails
# (common in some serverless/vendored environments), fall back to simple
# PostgREST HTTP calls using `requests`.
try:
    from supabase import create_client, Client  # type: ignore
    _HAS_SUPABASE_PY = True
except Exception:
    _HAS_SUPABASE_PY = False

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'your-secret-key-here-change-in-production')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)

# Supabase configuration
SUPABASE_URL = os.environ.get('SUPABASE_URL', "https://kflhmnyikwaznzkgeini.supabase.co")
SUPABASE_KEY = os.environ.get('SUPABASE_KEY', "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImtmbGhtbnlpa3dhem56a2dlaW5pIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjQ1NTQ5MDEsImV4cCI6MjA4MDEzMDkwMX0.nTkrtt4SQEYLq529ccYvnR46L1mbzBzagoVifP2VkWQ")
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_EMAIL = os.environ.get('SMTP_EMAIL', 'pmcplacements@gmail.com')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', 'xovepwitamhpmiac')
OTP_EXPIRY_MINUTES = int(os.environ.get('OTP_EXPIRY_MINUTES', '10'))


class _TableWrapper:
    def __init__(self, client, table_name):
        self._client = client
        self._table = table_name
        self._select = '*'
        self._filters = []
        self._payload = None
        self._extra_select_params = {}

    def select(self, cols='*', **kwargs):
        self._select = cols
        self._extra_select_params = kwargs or {}
        return self

    def eq(self, col, val):
        self._filters.append((col, val))
        return self

    def insert(self, payload):
        self._payload = payload
        self._op = 'insert'
        return self

    def upsert(self, payload):
        self._payload = payload
        self._op = 'upsert'
        return self

    def execute(self):
        # If we have a real supabase client, delegate
        if _HAS_SUPABASE_PY and getattr(self._client, 'table', None):
            try:
                t = self._client.table(self._table)
                if getattr(self, '_select', None):
                    try:
                        t = t.select(self._select, **(self._extra_select_params or {}))
                    except Exception:
                        t = t.select(self._select)
                for col, val in self._filters:
                    t = t.eq(col, val)
                if getattr(self, '_op', None) == 'insert':
                    t = t.insert(self._payload)
                elif getattr(self, '_op', None) == 'upsert':
                    t = t.upsert(self._payload)
                return t.execute()
            except Exception:
                # Fall through to HTTP fallback
                pass

        # Build headers for PostgREST
        headers = {
            'apikey': SUPABASE_KEY,
            'Authorization': f'Bearer {SUPABASE_KEY}',
            'Content-Type': 'application/json'
        }

        base = SUPABASE_URL.rstrip('/') + f"/rest/v1/{self._table}"

        if getattr(self, '_op', None) in ('insert', 'upsert'):
            # Use POST for insert/upsert. For upsert, request merge-duplicates.
            prefer = 'return=representation'
            if getattr(self, '_op', None) == 'upsert':
                prefer = 'resolution=merge-duplicates,return=representation'
            headers['Prefer'] = prefer
            resp = requests.post(base, json=self._payload, headers=headers)
            class R: pass
            r = R()
            try:
                r.data = resp.json()
            except Exception:
                r.data = None
            r.error = None if resp.ok else {'message': resp.text}
            return r

        # Default: select
        params = { 'select': self._select }
        # If client requested count via select(..., count='exact'), map to Prefer header
        prefer_count = None
        if self._extra_select_params and 'count' in self._extra_select_params:
            prefer_count = f"count={self._extra_select_params['count']}"
        for col, val in self._filters:
            # Use PostgREST eq operator
            params[f'{col}'] = f'eq.{val}'

        if prefer_count:
            headers['Prefer'] = prefer_count
        resp = requests.get(base, params=params, headers=headers)
        class R: pass
        r = R()
        try:
            r.data = resp.json()
        except Exception:
            r.data = None
        r.error = None if resp.ok else {'message': resp.text}
        return r


class SupabaseCompat:
    def __init__(self, url, key):
        self.url = url
        self.key = key
        if _HAS_SUPABASE_PY:
            try:
                self._client = create_client(url, key)
            except Exception:
                self._client = None
        else:
            self._client = None

    def table(self, name):
        return _TableWrapper(self._client, name)


# Create a compat supabase object (uses env vars if present)
supabase = SupabaseCompat(SUPABASE_URL, SUPABASE_KEY)

# Authentication decorator
def login_required(role=None):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'email' not in session:
                return redirect(url_for('index'))
            if role and session.get('role') != role:
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def _first_record(data):
    if isinstance(data, list):
        return data[0] if data else None
    if isinstance(data, dict):
        return data
    return None


def _find_account_by_email(role, email):
    email = (email or '').strip().lower()
    if role == 'admin':
        resp = supabase.table('users').select('*').eq('email', email).eq('role', 'admin').execute()
        return _first_record(resp.data), 'users'
    if role == 'student':
        resp = supabase.table('students').select('*').eq('email', email).execute()
        return _first_record(resp.data), 'students'
    return None, None


def _send_reset_otp_email(recipient_email, recipient_name, otp_code, role):
    msg = EmailMessage()
    msg['Subject'] = f"Placement Cell {role.title()} Password Reset OTP"
    msg['From'] = SMTP_EMAIL
    msg['To'] = recipient_email
    msg.set_content(
        f"""
Hello {recipient_name or 'User'},

We received a request to reset your {role} account password.

Your OTP is: {otp_code}
This OTP expires in {OTP_EXPIRY_MINUTES} minutes.

If you did not request this reset, please ignore this email.

Regards,
Placement Cell Team
""".strip()
    )

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as server:
        server.starttls()
        server.login(SMTP_EMAIL, SMTP_PASSWORD)
        server.send_message(msg)


def _update_password_by_email(role, email, new_password):
    table = 'users' if role == 'admin' else 'students'
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=representation'
    }

    params = {'email': f'eq.{email.strip().lower()}'}
    if role == 'admin':
        params['role'] = 'eq.admin'

    url = SUPABASE_URL.rstrip('/') + f"/rest/v1/{table}"
    resp = requests.patch(url, params=params, json={'password': new_password}, headers=headers)
    if not resp.ok:
        return {'ok': False, 'message': resp.text}
    return {'ok': True}

# Routes
@app.route('/')
def index():
    active_role = (request.args.get('role') or 'student').strip().lower()
    if active_role not in ('student', 'admin'):
        active_role = 'student'
    return render_template('index.html', active_role=active_role)


@app.route('/media/<path:filename>')
def media_file(filename):
    return send_from_directory(os.path.join(app.root_path, 'media'), filename)

@app.route('/student_login', methods=['GET', 'POST'])
def student_login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        source = (request.form.get('source') or '').strip().lower()
        
        try:
            response = supabase.table('students').select('*').eq('email', email).eq('password', password).execute()
            if response.data and len(response.data) > 0:
                student = response.data[0]
                session['email'] = student['email']
                session['name'] = student['name']
                session['student_id'] = student['id']
                session['role'] = 'student'
                session.permanent = True
                return redirect(url_for('student_dashboard'))
            else:
                if source == 'index':
                    return render_template('index.html', active_role='student', student_error='Invalid credentials')
                return render_template('student_login.html', error='Invalid credentials')
        except Exception as e:
            if source == 'index':
                return render_template('index.html', active_role='student', student_error=str(e))
            return render_template('student_login.html', error=str(e))
    
    return render_template('student_login.html')

@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        source = (request.form.get('source') or '').strip().lower()
        
        try:
            response = supabase.table('users').select('*').eq('email', email).eq('password', password).eq('role', 'admin').execute()
            if response.data and len(response.data) > 0:
                admin = response.data[0]
                session['email'] = admin['email']
                session['name'] = admin.get('name', 'Admin')
                session['user_id'] = admin['id']
                session['role'] = 'admin'
                session.permanent = True
                return redirect(url_for('admin_dashboard'))
            else:
                if source == 'index':
                    return render_template('index.html', active_role='admin', admin_error='Invalid credentials')
                return render_template('admin_login.html', error='Invalid credentials')
        except Exception as e:
            if source == 'index':
                return render_template('index.html', active_role='admin', admin_error=str(e))
            return render_template('admin_login.html', error=str(e))
    
    return render_template('admin_login.html')


@app.route('/forgot_password/<role>', methods=['GET', 'POST'])
def forgot_password(role):
    role = (role or '').strip().lower()
    if role not in ('admin', 'student'):
        return redirect(url_for('index'))

    if request.method == 'POST':
        email = (request.form.get('email') or '').strip().lower()
        account, _table = _find_account_by_email(role, email)

        if not account:
            return render_template('forgot_password.html', role=role, error='Email not found for this account type.')

        otp_code = f"{secrets.randbelow(1000000):06d}"
        otp_hash = hashlib.sha256(otp_code.encode('utf-8')).hexdigest()
        expires_at = int(time.time()) + (OTP_EXPIRY_MINUTES * 60)

        session['password_reset'] = {
            'role': role,
            'email': email,
            'otp_hash': otp_hash,
            'expires_at': expires_at
        }

        try:
            _send_reset_otp_email(email, account.get('name', 'User'), otp_code, role)
        except Exception as e:
            return render_template('forgot_password.html', role=role, error=f'Failed to send OTP email: {e}')

        return redirect(url_for('reset_password', role=role))

    return render_template('forgot_password.html', role=role)


@app.route('/reset_password/<role>', methods=['GET', 'POST'])
def reset_password(role):
    role = (role or '').strip().lower()
    if role not in ('admin', 'student'):
        return redirect(url_for('index'))

    reset_ctx = session.get('password_reset') or {}
    if not reset_ctx or reset_ctx.get('role') != role:
        return redirect(url_for('forgot_password', role=role))

    if request.method == 'POST':
        otp = (request.form.get('otp') or '').strip()
        new_password = (request.form.get('new_password') or '').strip()
        confirm_password = (request.form.get('confirm_password') or '').strip()

        if len(otp) != 6 or not otp.isdigit():
            return render_template('reset_password.html', role=role, error='Enter a valid 6-digit OTP.')

        if len(new_password) < 4:
            return render_template('reset_password.html', role=role, error='Password must be at least 4 characters.')

        if new_password != confirm_password:
            return render_template('reset_password.html', role=role, error='Passwords do not match.')

        if int(time.time()) > int(reset_ctx.get('expires_at', 0)):
            session.pop('password_reset', None)
            return render_template('forgot_password.html', role=role, error='OTP expired. Request a new OTP.')

        incoming_hash = hashlib.sha256(otp.encode('utf-8')).hexdigest()
        if incoming_hash != reset_ctx.get('otp_hash'):
            return render_template('reset_password.html', role=role, error='Invalid OTP.')

        email = reset_ctx.get('email', '').strip().lower()
        account, _table = _find_account_by_email(role, email)
        if not account:
            session.pop('password_reset', None)
            return render_template('forgot_password.html', role=role, error='Account no longer exists. Contact administrator.')

        update_result = _update_password_by_email(role, email, new_password)
        if not update_result.get('ok'):
            return render_template('reset_password.html', role=role, error=f"Unable to update password: {update_result.get('message')}")

        session.pop('password_reset', None)
        if role == 'admin':
            return redirect(url_for('index', role='admin', reset='success'))
        return redirect(url_for('index', role='student', reset='success'))

    return render_template('reset_password.html', role=role)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/student_dashboard')
@login_required(role='student')
def student_dashboard():
    return render_template('student_dashboard.html')

@app.route('/admin_dashboard')
@login_required(role='admin')
def admin_dashboard():
    return render_template('admin_dashboard.html')

@app.route('/exam_instructions')
@login_required(role='student')
def exam_instructions():
    exam_id = request.args.get('exam_id')
    return render_template('exam_instructions.html', exam_id=exam_id)

@app.route('/exam_fullscreen')
@login_required(role='student')
def exam_fullscreen():
    return render_template('exam_fullscreen.html')

@app.route('/student_management')
@login_required(role='admin')
def student_management():
    return render_template('student_management.html')

@app.route('/user_management')
@login_required(role='admin')
def user_management():
    return render_template('user_management.html', current_email=session.get('email', ''))

@app.route('/events_management')
@login_required(role='admin')
def events_management():
    return render_template('events_management.html')

@app.route('/circulars_management')
@login_required(role='admin')
def circulars_management():
    return render_template('circulars_management.html')

@app.route('/manage_exams')
@login_required(role='admin')
def manage_exams():
    return render_template('manage_exams.html')

@app.route('/create_exam')
@login_required(role='admin')
def create_exam():
    return render_template('create_exam.html')

@app.route('/question_bank')
@login_required(role='admin')
def question_bank():
    return render_template('question_bank.html')

@app.route('/results_analytics')
@login_required(role='admin')
def results_analytics():
    return render_template('results_analytics.html')

@app.route('/live_monitoring')
@login_required(role='admin')
def live_monitoring():
    return render_template('live_monitoring.html')

# API endpoints for AJAX calls
@app.route('/api/get_exams', methods=['GET'])
@login_required(role='student')
def api_get_exams():
    try:
        student_id = session.get('student_id')
        if not student_id:
            return jsonify({'error': 'Student not logged in'}), 401
            
        # Get assigned exams
        assigned = supabase.table('student_exam_map').select('exam_id').eq('student_id', student_id).execute()
        
        # Get submissions (removed 'status' field that doesn't exist)
        submissions = supabase.table('submissions').select('exam_id,score,total,percentage,submitted_at,auto_submitted').eq('student_id', student_id).execute()
        
        exam_list = []
        submitted_map = {sub['exam_id']: sub for sub in (submissions.data or [])}
        
        for map_item in (assigned.data or []):
            try:
                exam = supabase.table('exams').select('*').eq('id', map_item['exam_id']).execute()
                if exam.data and len(exam.data) > 0:
                    exam_data = exam.data[0]
                    exam_id = exam_data.get('id')
                    exam_data['submitted'] = exam_id in submitted_map
                    exam_data['submission'] = submitted_map.get(exam_id)
                    exam_list.append(exam_data)
            except Exception as exam_error:
                print(f"Error loading exam: {exam_error}")
                continue
        
        return jsonify({'exams': exam_list})
    except Exception as e:
        print(f"Error in get_exams: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/get_results', methods=['GET'])
@login_required(role='student')
def api_get_results():
    try:
        student_id = session.get('student_id')
        if not student_id:
            return jsonify({'error': 'Student not logged in'}), 401
            
        submissions = supabase.table('submissions').select('*').eq('student_id', student_id).execute()
        
        results = []
        for sub in (submissions.data or []):
            try:
                exam = supabase.table('exams').select('title').eq('id', sub.get('exam_id')).execute()
                
                # Calculate percentage
                score = float(sub.get('score', 0))
                total = float(sub.get('total', 1))
                percentage = (score / total * 100) if total > 0 else 0
                
                # Determine status based on submission
                status = 'Auto-Submitted' if sub.get('auto_submitted') else 'Completed'
                
                result = {
                    'exam_title': exam.data[0]['title'] if exam.data and len(exam.data) > 0 else 'Unknown Exam',
                    'score': round(percentage, 2),
                    'raw_score': f"{int(score)}/{int(total)}",
                    'status': status,
                    'submitted_at': sub.get('submitted_at'),
                    'percentage': round(percentage, 2)
                }
                results.append(result)
            except Exception as sub_error:
                print(f"Error processing submission: {sub_error}")
                continue
        
        return jsonify({'results': results})
    except Exception as e:
        print(f"Error in get_results: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/submit_exam', methods=['POST'])
@login_required(role='student')
def api_submit_exam():
    try:
        data = request.json
        student_id = session.get('student_id')
        exam_id = data.get('exam_id')
        answers = data.get('answers')
        score = data.get('score')
        
        # Store in submissions table
        submission_obj = {
            'student_id': student_id,
            'exam_id': exam_id,
            'answers': str(answers),
            'score': score,
            'status': 'submitted'
        }
        
        supabase.table('submissions').insert(submission_obj).execute()
        supabase.table('results').upsert({'student_id': student_id, 'exam_id': exam_id, 'score': score}).execute()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/get_questions', methods=['GET'])
@login_required(role='student')
def api_get_questions():
    try:
        exam_id = request.args.get('exam_id')
        questions = supabase.table('questions').select('*').eq('exam_id', exam_id).execute()
        return jsonify({'questions': questions.data or []})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/save_answer', methods=['POST'])
@login_required(role='student')
def api_save_answer():
    try:
        data = request.json
        student_id = session.get('student_id')
        question_id = data.get('question_id')
        answer = data.get('answer')
        
        supabase.table('answers').upsert({
            'student_id': student_id,
            'question_id': question_id,
            'answer': answer
        }).execute()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/get_student_profile', methods=['GET'])
@login_required(role='student')
def api_get_student_profile():
    try:
        email = session.get('email')
        student = supabase.table('students').select('*').eq('email', email).execute()
        if student.data:
            return jsonify({'student': student.data[0]})
        return jsonify({'error': 'Student not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload_word_questions', methods=['POST'])
@login_required(role='admin')
def api_upload_word_questions():
    try:
        import re
        from docx import Document
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        exam_id = request.form.get('exam_id')
        section = request.form.get('section', '')
        
        if not exam_id:
            return jsonify({'error': 'No exam ID provided'}), 400
        
        # Parse Word document
        doc = Document(file)
        questions = []
        current_question = {}
        
        for para in doc.paragraphs:
            text = para.text.strip()
            if not text:
                continue
            
            # Format 1: Q1. Question? or Question: text
            if re.match(r'^(Q\d+\.|Question:)', text, re.IGNORECASE):
                if current_question and 'question_text' in current_question:
                    questions.append(current_question)
                current_question = {'question_text': re.sub(r'^(Q\d+\.|Question:)\s*', '', text, flags=re.IGNORECASE)}
                current_question['options'] = []
            
            # Format 1: Options A) B) C) D)
            elif re.match(r'^[A-D][\)\.]\s*.+', text, re.IGNORECASE):
                option_text = re.sub(r'^[A-D][\)\.]\s*', '', text, flags=re.IGNORECASE)
                is_correct = '[CORRECT]' in option_text.upper() or '(CORRECT)' in option_text.upper()
                option_text = re.sub(r'\s*[\[\(]CORRECT[\]\)]\s*', '', option_text, flags=re.IGNORECASE).strip()
                current_question['options'].append(option_text)
                if is_correct:
                    current_question['correct_answer'] = option_text
            
            # Format 2: Option1:, Option2:, etc.
            elif re.match(r'^Option[1-4]:\s*.+', text, re.IGNORECASE):
                option_text = re.sub(r'^Option[1-4]:\s*', '', text, flags=re.IGNORECASE).strip()
                current_question['options'].append(option_text)
            
            # Format 2: Correct: Option2
            elif re.match(r'^Correct:\s*.+', text, re.IGNORECASE):
                correct_ref = re.sub(r'^Correct:\s*', '', text, flags=re.IGNORECASE).strip()
                if 'Option' in correct_ref:
                    opt_num = int(re.search(r'\d+', correct_ref).group()) - 1
                    if opt_num < len(current_question.get('options', [])):
                        current_question['correct_answer'] = current_question['options'][opt_num]
                else:
                    current_question['correct_answer'] = correct_ref
            
            # Difficulty
            elif re.match(r'^Difficulty:\s*.+', text, re.IGNORECASE):
                current_question['difficulty'] = re.sub(r'^Difficulty:\s*', '', text, flags=re.IGNORECASE).strip()
            
            # Tags
            elif re.match(r'^Tags:\s*.+', text, re.IGNORECASE):
                current_question['tags'] = re.sub(r'^Tags:\s*', '', text, flags=re.IGNORECASE).strip()
            
            # Negative Mark
            elif re.match(r'^Negative:\s*.+', text, re.IGNORECASE):
                try:
                    current_question['negative_mark'] = float(re.sub(r'^Negative:\s*', '', text, flags=re.IGNORECASE).strip())
                except:
                    pass
        
        # Add last question
        if current_question and 'question_text' in current_question:
            questions.append(current_question)
        
        # Insert questions into database
        success_count = 0
        failed_count = 0
        preview = []
        
        for q in questions:
            if 'question_text' not in q or len(q.get('options', [])) < 4 or 'correct_answer' not in q:
                failed_count += 1
                continue
            
            insert_data = {
                'exam_id': exam_id,
                'question_text': q['question_text'],
                'type': 'mcq',
                'options': q['options'],  # Supabase will convert to JSON
                'correct_answer': q['correct_answer']
            }
            
            if section:
                insert_data['section'] = section
            if 'difficulty' in q:
                insert_data['difficulty'] = q['difficulty']
            if 'tags' in q:
                insert_data['tags'] = q['tags']
            if 'negative_mark' in q:
                insert_data['negative_mark'] = q['negative_mark']
            
            try:
                result = supabase.table('questions').insert(insert_data).execute()
                if result.data:
                    success_count += 1
                    if len(preview) < 3:
                        preview.append({'question_text': q['question_text']})
                else:
                    failed_count += 1
            except Exception as e:
                print(f"Error inserting question: {e}")
                failed_count += 1
        
        return jsonify({
            'success': success_count,
            'failed': failed_count,
            'preview': preview
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard_stats', methods=['GET'])
@login_required(role='admin')
def dashboard_stats():
    try:
        # Get total students count
        students_response = supabase.table('students').select('id', count='exact').execute()
        total_students = students_response.count if hasattr(students_response, 'count') else len(students_response.data)
        
        # Get total exams count
        exams_response = supabase.table('exams').select('id', count='exact').execute()
        total_exams = exams_response.count if hasattr(exams_response, 'count') else len(exams_response.data)
        
        # Get total questions count
        questions_response = supabase.table('questions').select('id', count='exact').execute()
        total_questions = questions_response.count if hasattr(questions_response, 'count') else len(questions_response.data)
        
        # Get total submissions count
        submissions_response = supabase.table('submissions').select('id', count='exact').execute()
        total_submissions = submissions_response.count if hasattr(submissions_response, 'count') else len(submissions_response.data)
        
        return jsonify({
            'students': total_students,
            'exams': total_exams,
            'questions': total_questions,
            'submissions': total_submissions
        })
    except Exception as e:
        return jsonify({
            'error': str(e),
            'students': 0,
            'exams': 0,
            'questions': 0,
            'submissions': 0
        }), 200

@app.route('/api/exam_excel_report', methods=['GET'])
@login_required(role='admin')
def api_exam_excel_report():
    try:
        import io
        import json as _json
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from flask import send_file

        exam_id = request.args.get('exam_id')
        if not exam_id:
            return jsonify({'error': 'exam_id is required'}), 400

        # Exam details
        exam_resp = supabase.table('exams').select('*').eq('id', exam_id).execute()
        if not exam_resp.data:
            return jsonify({'error': 'Exam not found'}), 404
        exam = exam_resp.data[0]

        # Questions for this exam
        q_resp = supabase.table('questions').select('id,question_text,correct_answer,type').eq('exam_id', exam_id).execute()
        questions = q_resp.data or []

        # Assigned students
        assigned_resp = supabase.table('student_exam_map').select('student_id').eq('exam_id', exam_id).execute()
        assigned_ids = list(dict.fromkeys([str(r['student_id']) for r in (assigned_resp.data or [])]))  # Deduplicate preserving order

        # Submissions
        subs_resp = supabase.table('submissions').select('*').eq('exam_id', exam_id).execute()
        subs_by_student = {str(s['student_id']): s for s in (subs_resp.data or [])}

        # Fallback: if no assignment mapping, use submitted student IDs
        if not assigned_ids:
            assigned_ids = list(subs_by_student.keys())

        # Student details
        students_map = {}
        for sid in assigned_ids:
            s_resp = supabase.table('students').select('*').eq('id', sid).execute()
            if s_resp.data:
                # Handle both list and dict responses
                if isinstance(s_resp.data, list):
                    students_map[sid] = s_resp.data[0]
                else:
                    students_map[sid] = s_resp.data

        # Sort: submitted first, then by sid
        submitted_ids = set(subs_by_student.keys())
        sorted_ids = sorted(assigned_ids, key=lambda s: (0 if s in submitted_ids else 1, s))

        # ── Build workbook ───────────────────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = 'Student Report'

        # Styles
        h_blue   = PatternFill('solid', fgColor='1A3C6E')
        h_green  = PatternFill('solid', fgColor='1E7145')
        g_green  = PatternFill('solid', fgColor='C6EFCE')
        g_red    = PatternFill('solid', fgColor='FFC7CE')
        g_yellow = PatternFill('solid', fgColor='FFEB9C')
        g_gray   = PatternFill('solid', fgColor='E0E0E0')
        w_bold   = Font(bold=True, color='FFFFFF')
        center   = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Row 1 — Exam title banner
        ncols = 7 + len(questions)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ncols, 8))
        tc = ws.cell(row=1, column=1, value=f"Student Report \u2014 {exam.get('title', exam_id)}")
        tc.font = Font(bold=True, size=13, color='FFFFFF')
        tc.fill = h_blue
        tc.alignment = center
        ws.row_dimensions[1].height = 24

        # Row 2 — Column headers
        static_cols = ['#', 'Student Name', 'Email', 'Department', 'Roll No', 'Score (%)', 'Status']
        q_labels    = [f'Q{i+1}' for i in range(len(questions))]
        all_cols    = static_cols + q_labels
        for ci, hdr in enumerate(all_cols, start=1):
            cell = ws.cell(row=2, column=ci, value=hdr)
            cell.font  = w_bold
            cell.fill  = h_blue if ci <= len(static_cols) else h_green
            cell.alignment = center
        ws.row_dimensions[2].height = 18

        # Row 3 — Question sub-headers (short text)
        for ci, q in enumerate(questions, start=len(static_cols) + 1):
            txt = q['question_text']
            cell = ws.cell(row=3, column=ci, value=(txt[:70] + '\u2026' if len(txt) > 70 else txt))
            cell.font = Font(italic=True, size=8, color='444444')
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.row_dimensions[3].height = 42

        # Data rows (start at row 4)
        for row_num, sid in enumerate(sorted_ids, start=4):
            student = students_map.get(sid, {})
            sub     = subs_by_student.get(sid)

            name  = student.get('name', 'Unknown')
            email = student.get('email', '')
            dept  = student.get('department', student.get('dept', ''))
            roll  = student.get('roll_no', student.get('roll_number', student.get('usn', '')))

            if sub:
                try:
                    score_pct = round(float(sub.get('percentage', sub.get('score', 0))), 2)
                except Exception:
                    score_pct = 0
                status = 'Auto-Submitted' if sub.get('auto_submitted') else 'Submitted'
            else:
                score_pct = None
                status    = 'Not Submitted'

            ws.cell(row=row_num, column=1, value=row_num - 3).alignment = center
            ws.cell(row=row_num, column=2, value=name)
            ws.cell(row=row_num, column=3, value=email)
            ws.cell(row=row_num, column=4, value=dept)
            ws.cell(row=row_num, column=5, value=roll)

            sc = ws.cell(row=row_num, column=6, value=score_pct if score_pct is not None else 'N/A')
            sc.alignment = center
            if score_pct is not None:
                sc.fill = g_green if score_pct >= 60 else g_red

            stc = ws.cell(row=row_num, column=7, value=status)
            stc.alignment = center
            stc.fill = (g_green if status == 'Submitted'
                        else g_yellow if status == 'Auto-Submitted'
                        else g_gray)

            # Per-question columns
            answers = {}
            if sub and sub.get('answers'):
                try:
                    a = sub['answers']
                    answers = a if isinstance(a, dict) else _json.loads(a)
                except Exception:
                    answers = {}

            for qi, q in enumerate(questions):
                col = len(static_cols) + qi + 1
                qid = str(q['id'])
                s_ans = answers.get(qid, answers.get(q['id']))
                c_ans = q.get('correct_answer', '')

                if sub is None:
                    val, fill = 'N/A', g_gray
                elif s_ans is None:
                    val, fill = 'Not Answered', g_gray
                elif str(s_ans).strip().upper() == str(c_ans).strip().upper():
                    val, fill = 'Correct', g_green
                else:
                    val, fill = 'Wrong', g_red

                cell = ws.cell(row=row_num, column=col, value=val)
                cell.fill = fill
                cell.alignment = center

        # Column widths
        col_widths = [4, 22, 28, 16, 14, 10, 14] + [12] * len(questions)
        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.freeze_panes = 'A4'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_title = exam.get('title', 'exam').replace(' ', '_')[:50]
        filename = f"{safe_title}_student_report.xlsx"
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    import os
    debug_env = os.environ.get('FLASK_DEBUG', 'False').lower() in ('1', 'true', 'yes')
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=debug_env, host='0.0.0.0', port=port)
